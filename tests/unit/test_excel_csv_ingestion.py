"""Excel and CSV ingestion via the parse_docs OCR pipeline."""

from io import BytesIO

import openpyxl

from wealthtax_agent.parse_docs import _infer_mime_type, ocr_bytes_to_text


def _make_xlsx_bytes(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schwab 1099-B"
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_mime_inference():
    xlsx_bytes = _make_xlsx_bytes([["Header"], [1]])
    mime = _infer_mime_type(xlsx_bytes, None)
    assert mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_csv_mime_inference():
    csv_bytes = b"col1,col2,col3\n1,2,3\n4,5,6\n"
    mime = _infer_mime_type(csv_bytes, None)
    assert mime == "text/csv"


def test_pdf_mime_inference_takes_priority():
    pdf_bytes = b"%PDF-1.4\n..."
    mime = _infer_mime_type(pdf_bytes, None)
    assert mime == "application/pdf"


def test_xlsx_extracted_to_tab_separated_text():
    xlsx_bytes = _make_xlsx_bytes([
        ["Symbol", "Proceeds", "Cost Basis", "Gain/Loss", "Term"],
        ["AAPL", 12000, 8000, 4000, "LT"],
        ["TSLA", 5000, 7000, -2000, "ST"],
    ])
    text = ocr_bytes_to_text(xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "Schwab 1099-B" in text
    assert "AAPL\t12000\t8000\t4000\tLT" in text
    assert "TSLA\t5000\t7000\t-2000\tST" in text


def test_csv_extracted_to_newline_separated_text():
    csv_bytes = b"Symbol,Proceeds,CostBasis\nAAPL,12000,8000\nTSLA,5000,7000\n"
    text = ocr_bytes_to_text(csv_bytes, "text/csv")
    assert "Symbol,Proceeds,CostBasis" in text
    assert "AAPL,12000,8000" in text
    assert "TSLA,5000,7000" in text


def test_xlsx_handles_multiple_sheets():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Salary"
    ws1.append(["Gross", 1800000])
    ws2 = wb.create_sheet("Deductions")
    ws2.append(["80C", 150000])
    buf = BytesIO()
    wb.save(buf)
    text = ocr_bytes_to_text(buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "# Sheet: Salary" in text
    assert "# Sheet: Deductions" in text
    assert "1800000" in text
    assert "150000" in text
